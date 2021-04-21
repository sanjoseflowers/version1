from django.contrib.auth import login,logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.views import LoginView
from django.http import HttpResponseRedirect
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.views.generic import FormView, RedirectView, TemplateView
from django.contrib.auth.models import User
from django.db.models import Sum
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse


import configuraciones.settings as setting
from django.shortcuts import render
from base.pagina_web.models import Acerca, Galeria, Producto, Contacto, Categoria
from direccion_envio.models import Orden, DireccionEnvio
from carrito.models import Carrito, CarritoProductos
from django.views.generic import ListView, CreateView, DeleteView, UpdateView

from .forms import AcercaForm, GaleriaForm, ProductosForm, ContactosForm, PedidosForm, CategoriaForm
from django.contrib import messages
from .forms import RegisterForm

from openpyxl import Workbook
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side
# Create your views here.

class LoginFormView(LoginView):
    template_name = "usuarios/ingresar.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            return redirect(setting.LOGIN_REDIRECT_URL)

        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = 'Iniciar sesión'
        return context
    

class LogoutView(RedirectView):
    pattern_name = 'ingresar'

    def dispatch(self, request, *args, **kwargs):
        logout(request)
        return super().dispatch(request, *args, **kwargs)

def register_view(request):
    form = RegisterForm(request.POST or None)
    
    if request.method == 'POST' and form.is_valid():
        username = form.cleaned_data.get('username')
        email = form.cleaned_data.get('email')
        password = form.cleaned_data.get('password')

        user = User.objects.create_user(username, email, password)
        if user:
            login(request, user)
            messages.success(request, 'Usuario creado con éxito')
            return redirect('inicio')

    return render(request, 'usuarios/registrar.html', {'form': form})


def administrador(request):
    clientes = User.objects.all()
    productos = Producto.objects.all()
    ordenes = Orden.objects.all()
    aux_clientes = len(clientes)
    clientes_total = int(aux_clientes) - 1

    my_context = {
        'clientes_total':clientes_total,
        'productos':productos,
        'ordenes':ordenes,
        'total_ganancias': Carrito.objects.all().aggregate(Sum('total'))
    }
    return render(request, 'administrador/inicio.html', my_context)

class AcercaListView(LoginRequiredMixin, ListView):
    model = Acerca
    template_name = 'administrador/acerca.html'

class AcercaCrearView(LoginRequiredMixin, CreateView):
    model = Acerca
    form_class = AcercaForm
    template_name = 'administrador/acerca/crear.html'
    success_url = reverse_lazy('acerca_lista')
    url_redirect = success_url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Ingresar información acerca de la página'
        context['list_url'] = self.success_url
        context['action'] = 'add'
        return context

class AcercaActualizarView(LoginRequiredMixin, UpdateView):
    model = Acerca
    form_class = AcercaForm
    template_name = 'administrador/acerca/crear.html'
    success_url = reverse_lazy('acerca_lista')
    url_redirect = success_url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        id_data = self.kwargs.get('pk')
        context['title'] = 'Actualizar información acerca de la página'
        context['list_url'] = self.success_url
        context['action'] = 'edit'
        return context

class AcercaEliminarView(LoginRequiredMixin, DeleteView):
    model = Acerca
    form_class = AcercaForm
    template_name = 'administrador/acerca/eliminar.html'
    success_url = reverse_lazy('acerca_lista')
    url_redirect = success_url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Eliminación de acerca'
        context['list_url'] = self.success_url
        return context
#categoria
class CategoriaListView(LoginRequiredMixin, ListView):
    model = Categoria
    template_name = 'administrador/categoria.html'

class CategoriaCrearView(LoginRequiredMixin, CreateView):
    model = Categoria
    form_class = CategoriaForm
    template_name = 'administrador/categoria/crear.html'
    success_url = reverse_lazy('categoria_lista')
    url_redirect = success_url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Ingresar información categoria de la página'
        context['list_url'] = self.success_url
        context['action'] = 'add'
        return context
        
class CategoriaActualizarView(LoginRequiredMixin, UpdateView):
    model = Categoria
    form_class = CategoriaForm
    template_name = 'administrador/categoria/crear.html'
    success_url = reverse_lazy('categoria_lista')
    url_redirect = success_url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        id_data = self.kwargs.get('pk')
        context['title'] = 'Actualizar información categoria de la página'
        context['list_url'] = self.success_url
        context['action'] = 'edit'
        return context

class CategoriaEliminarView(LoginRequiredMixin, DeleteView):
    model = Categoria
    form_class = CategoriaForm
    template_name = 'administrador/categoria/eliminar.html'
    success_url = reverse_lazy('categoria_lista')
    url_redirect = success_url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Eliminación de categoria'
        context['list_url'] = self.success_url
        return context

#galeria
class GaleriaListView(LoginRequiredMixin, ListView):
    model = Galeria
    template_name = 'administrador/galeria.html'

class GaleriaCrearView(LoginRequiredMixin, CreateView):
    model = Galeria
    form_class = GaleriaForm
    template_name = 'administrador/galeria/crear.html'
    success_url = reverse_lazy('galeria_lista')
    url_redirect = success_url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Ingresar información galeria de la página'
        context['list_url'] = self.success_url
        context['action'] = 'add'
        return context
        
class GaleriaActualizarView(LoginRequiredMixin, UpdateView):
    model = Galeria
    form_class = GaleriaForm
    template_name = 'administrador/galeria/crear.html'
    success_url = reverse_lazy('galeria_lista')
    url_redirect = success_url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        id_data = self.kwargs.get('pk')
        context['title'] = 'Actualizar información galeria de la página'
        context['list_url'] = self.success_url
        context['action'] = 'edit'
        return context

class GaleriaEliminarView(LoginRequiredMixin, DeleteView):
    model = Galeria
    form_class = GaleriaForm
    template_name = 'administrador/galeria/eliminar.html'
    success_url = reverse_lazy('galeria_lista')
    url_redirect = success_url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Eliminación de galeria'
        context['list_url'] = self.success_url
        return context

#productos
class ProductosListView(LoginRequiredMixin, ListView):
    model = Producto
    template_name = 'administrador/productos.html'

class ProductosCrearView(LoginRequiredMixin, CreateView):
    model = Producto
    form_class = ProductosForm
    template_name = 'administrador/productos/crear.html'
    success_url = reverse_lazy('productos_lista')
    url_redirect = success_url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Ingresar información de productos de la página'
        context['list_url'] = self.success_url
        context['action'] = 'add'
        return context
        
class ProductosActualizarView(LoginRequiredMixin, UpdateView):
    model = Producto
    form_class = ProductosForm
    template_name = 'administrador/productos/editar.html'
    success_url = reverse_lazy('productos_lista')
    url_redirect = success_url

    def get_context_data(self, **kwargs):
        context = super(ProductosActualizarView, self).get_context_data(**kwargs)
        id_data = self.kwargs.get('pk')
        if 'form' not in context:
            context['form'] = self.form_class(self.request.POST, self.request.FILES,instance=Producto)
        context['title'] = 'Actualizar información galeria de la página'
        context['list_url'] = self.success_url
        context['action'] = 'edit'
        context['object_list'] = self.form_class(self.request.POST, self.request.FILES, instance=Producto)
        return context

class ProductosEliminarView(LoginRequiredMixin, DeleteView):
    model = Producto
    form_class = ProductosForm
    template_name = 'administrador/productos/eliminar.html'
    success_url = reverse_lazy('productos_lista')
    url_redirect = success_url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Eliminación de producto'
        context['list_url'] = self.success_url
        return context

#contactos
class ContactoListView(LoginRequiredMixin, ListView):
    model = Contacto
    template_name = 'administrador/contactos.html'

class ContactosCrearView(LoginRequiredMixin, CreateView):
    model = Contacto
    form_class = ContactosForm
    template_name = 'administrador/contacto/crear.html'
    success_url = reverse_lazy('contacto_lista')
    url_redirect = success_url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Ingresar contáctos de la página'
        context['list_url'] = self.success_url
        context['action'] = 'add'
        return context
        
class ContactosActualizarView(LoginRequiredMixin, UpdateView):
    model = Contacto
    form_class = ContactosForm
    template_name = 'administrador/contacto/crear.html'
    success_url = reverse_lazy('contacto_lista')
    url_redirect = success_url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        id_data = self.kwargs.get('pk')
        context['title'] = 'Actualizar contáctos de la página'
        context['list_url'] = self.success_url
        context['action'] = 'edit'
        return context

class ContactosEliminarView(LoginRequiredMixin, DeleteView):
    model = Contacto
    form_class = ContactosForm
    template_name = 'administrador/contacto/eliminar.html'
    success_url = reverse_lazy('contacto_lista')
    url_redirect = success_url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Eliminación de contacto'
        context['list_url'] = self.success_url
        return context

#pedidos
class PedidoListView(LoginRequiredMixin, ListView):
    model = Orden
    template_name = 'administrador/pedidos.html'

    def get_context_data(self, **kwargs):
        context = super(PedidoListView, self).get_context_data(**kwargs)
        context['action'] = 'edit'
        return context

class PedidosActualizarView(LoginRequiredMixin, UpdateView):
    model = Orden
    form_class = PedidosForm
    template_name = 'administrador/pedidos/todos.html'
    success_url = reverse_lazy('pedido_lista')
    url_redirect = success_url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        id_data = self.kwargs.get('pk')
        context['title'] = 'Despachar pedidos'
        context['list_url'] = self.success_url
        context['action'] = 'edit'
        context['object_list'] = Orden.objects.filter(pk=id_data)
        consulta = Orden.objects.filter(pk=id_data).first()
        context['items'] = CarritoProductos.objects.filter(carrito_id=consulta.carrito.id)
        context['direccion'] = DireccionEnvio.objects.filter(id=id_data)
        return context

class PedidosEliminarView(LoginRequiredMixin, DeleteView):
    model = Orden
    form_class = PedidosForm
    template_name = 'administrador/contacto/eliminar.html'
    success_url = reverse_lazy('pedido_lista')
    url_redirect = success_url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Eliminación de pedido'
        context['list_url'] = self.success_url
        return context


def reportes(request):
    ordenes = Orden.objects.all()
    productos = Producto.objects.all()
    
    wb = Workbook()
    ws = wb.active
    # Datos encuestador
    ws.title = "Reporte de ventas"
    ws["B1"].font = Font(size = 16, bold= True)
    ws["B1"] = "Ventas"
    ws["B3"].font = Font(size = 13, bold= True)
    ws["B3"] = "Fecha"
    ws["C3"].font = Font(size = 13, bold= True)
    ws["C3"] = "No"
    ws["D3"].font = Font(size = 13, bold= True)
    ws["D3"] = "Pedido"
    ws["E3"].font = Font(size = 13, bold= True)
    ws["E3"] = "Dirección"
    ws["F3"].font = Font(size = 13, bold= True)
    ws["F3"] = "Estado"
    ws["G3"].font = Font(size = 13, bold= True)
    ws["G3"] = "Total"
    cont = 4

    for orden in ordenes:
        ws.cell(row=cont, column=2).value = str(orden.created_at)
        ws.cell(row=cont, column=3).value = orden.id
        ws.cell(row=cont, column=4).value = str(orden.carrito)
        ws.cell(row=cont, column=5).value = str(orden.direccion)
        ws.cell(row=cont, column=6).value = str(orden.estado)
        ws.cell(row=cont, column=7).value = str(orden.carrito.total)
        cont += 1
    
    nombre_archivo = "ReporteVentas.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    content = "attachment; filename = {0}".format(nombre_archivo)
    response["Content-Disposition"] = content
    wb.save(response)
    return response

    return render(request, 'administrador/pedidos.html')